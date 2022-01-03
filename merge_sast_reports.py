import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

old_xls_file = input("Enter Old Excel File Path :")
new_xls_file = input("Enter New Excel File Path :")

oufile_list = os.path.splitext(new_xls_file)
base_dir = os.path.dirname(new_xls_file)
os.mkdir(os.path.join(base_dir, "merged"))
outfile = os.path.join(base_dir, 'merged', os.path.basename(new_xls_file))


df_old = pd.read_excel(old_xls_file)
df_new = pd.read_excel(new_xls_file)

for i in range(len(df_old)):
    category = df_old.loc[i, "Category"]
    source_file_path = df_old.loc[i, "Source File Path"]
    source_file_no = df_old.loc[i, "Source Line No"]
    sink_file_path = df_old.loc[i, "Sink File Path"]
    sink_file_no = df_old.loc[i, "Sink Line No"]
    comment = df_old.loc[i, "Auditor Comment"]
    for j in range(len(df_new)):
        if category == df_new.loc[j, "Category"] and source_file_path == df_new.loc[j, "Sorce File Path"] and source_file_no == df_new.loc[j, "Sorce File No"] and sink_file_path == df_new.loc[j, "Sink File Path"] and sink_file_no == df_new.loc[j, "Sink File No"] and df_new.loc[j, "Status"] == "Open":
            df_new.loc[j, "Developer Comment"] = "Not an issue "+str(comment.split("\n")[-1])

merged_df = df_new

workbook = Workbook()
workbook.remove(workbook.active)
header_font = Font(
	name='Calibri',
	bold=True,
	color='FFFFFF'
)
centered_alignment = Alignment(horizontal='center')
wrapped_alignment = Alignment(vertical='top',wrap_text=False)
sast_columns = [
	('Category',40),
	('Kingdom',15),
	('Abstract',45),
	('Priority',15),
	('Source File Path',45),
	('Source Line No',15),
	('Sink File Path',45),
	('Sink Line No',15),
	('Status',15),
	('Auditor Comment',45),
	('Developer Comment',45)
]

worksheet = workbook.create_sheet(title='vulnerabilities',index=0)
fill = PatternFill(
	start_color='5FABE6',
	end_color='5FABE6',
	fill_type="solid"
)
row_num = 1
for col_num, (column_title, column_width) in enumerate(sast_columns, 1):
	cell = worksheet.cell(row=row_num, column=col_num)
	cell.value = column_title
	cell.font = header_font
	cell.alignment = centered_alignment
	cell.fill = fill
	column_letter = get_column_letter(col_num)
	column_dimensions = worksheet.column_dimensions[column_letter]
	column_dimensions.width = column_width

for i in range(len(merged_df['Category'])):
	row_num += 1
	row = [
		(merged_df.loc[i,'Category'],'Normal'),
		(merged_df.loc[i,'Kingdom'],'Normal'),
		(merged_df.loc[i,'Abstract'],'Normal'),
		(merged_df.loc[i,'Priority'],'Normal'),
		(merged_df.loc[i,'Source File Path'],'Normal'),
		(merged_df.loc[i,'Source Line No'],'Normal'),
		(merged_df.loc[i,'Sink File Path'],'Normal'),
		(merged_df.loc[i,'Sink Line No'],'Normal'),
		(merged_df.loc[i,'Status'],'Normal'),
		(merged_df.loc[i,'Auditor Comment'],'Normal'),
		(merged_df.loc[i,'Developer Comment'],'Normal'),
	]
	for col_num, (cell_value, cell_format) in enumerate(row, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = cell_value
		cell.style = cell_format
		cell.alignment = wrapped_alignment


worksheet.freeze_panes = worksheet['A2']
worksheet.sheet_properties.tabColor = '5FABE6'
workbook.save()
